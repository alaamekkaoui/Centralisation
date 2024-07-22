from flask import Flask, request, jsonify, redirect, url_for
from flask_mysqldb import MySQL
import requests
from pywinauto.application import Application
import time
from flask_cors import CORS
from netmiko import ConnectHandler
import re
from pythonping import ping
from MySQLdb.cursors import DictCursor
import matplotlib.pyplot as plt
import retrieve_config_terminal
import logging
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime

app = Flask(__name__)
app.config['SECRET_KEY'] = 'azerty'
CORS(app, resources={r"/*": {"origins": "*", "methods": ["GET", "POST"], "allow_headers": ["Content-Type", "Authorization"]}})

# Configure MySQL connection
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'dxc-remote'

mysql = MySQL(app)

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email')
    password = data.get('password')

    cur = mysql.connection.cursor(DictCursor)
    cur.execute("SELECT * FROM users WHERE email = %s AND password = %s", (email, password))
    user = cur.fetchone()
    cur.close()

    if user:
        return jsonify({'message': 1}), 200
    else:
        return jsonify({'message': 0}), 200

@app.route('/logout', methods=['POST'])
def logout():
    # Add logic for logout, for example using Flask-Login
    return jsonify({'message': 'Logout successful'}), 200

@app.route('/client1', methods=['GET'])
def get_client1():
    cur = mysql.connection.cursor(DictCursor)
    cur.execute("SELECT * FROM client1")
    data = cur.fetchall()
    cur.close()
    return jsonify(data)

@app.route('/client2', methods=['GET'])
def get_client2():
    cur = mysql.connection.cursor(DictCursor)
    cur.execute("SELECT * FROM client2")
    data = cur.fetchall()
    cur.close()
    return jsonify(data)

@app.route('/home', methods=['GET'])
def home():
    cur = mysql.connection.cursor(DictCursor)
    cur.execute("SELECT * FROM client1")
    client1 = cur.fetchall()

    cur.execute("SELECT * FROM client2")
    client2 = cur.fetchall()

    cur.execute("SELECT COUNT(*) FROM client1")
    count_client1 = cur.fetchone()['COUNT(*)']

    cur.execute("SELECT COUNT(*) FROM client2")
    count_client2 = cur.fetchone()['COUNT(*)']
    cur.close()

    return jsonify({
        'client1': client1,
        'client2': client2,
        'client1Count': count_client1,
        'client2Count': count_client2
    })

def ping_devices():
    """
    Effectue des pings sur les adresses IP des dispositifs stockées dans les tables client1 et client2.
    Met à jour le statut des dispositifs dans la base de données en fonction du résultat du ping.
    """
    while True:
        try:
            cur = mysql.connection.cursor(DictCursor)
            # Récupérer les adresses IP des dispositifs depuis la table client1
            cur.execute("SELECT id, ip FROM client1")
            client1_devices = cur.fetchall()

            # Effectuer des pings sur les dispositifs de la table client1
            for device in client1_devices:
                device_id = device['id']
                ip_address = device['ip']
                response = ping(ip_address, count=1, timeout=1)
                status = "Connecté" if response.success() else "Déconnecté"
                cur.execute("UPDATE client1 SET status = %s WHERE id = %s", (status, device_id))

            # Récupérer les adresses IP des dispositifs depuis la table client2
            cur.execute("SELECT id, ip FROM client2")
            client2_devices = cur.fetchall()

            # Effectuer des pings sur les dispositifs de la table client2
            for device in client2_devices:
                device_id = device['id']
                ip_address = device['ip']
                response = ping(ip_address, count=1, timeout=1)
                status = "Connecté" if response.success() else "Déconnecté"
                cur.execute("UPDATE client2 SET status = %s WHERE id = %s", (status, device_id))

            mysql.connection.commit()
            cur.close()

            # Attendre 30 secondes avant de refaire un ping
            time.sleep(30)

        except Exception as e:
            print(f"Erreur MySQL: {e}")
            # Vous pouvez ajouter d'autres actions de gestion d'erreur ici, comme une pause et une nouvelle tentative de connexion

# Appeler la fonction pour effectuer des pings sur les dispositifs et mettre à jour leurs statuts dans la base de données
@app.route('/ping_devices')
def ping_devices_view():
    ping_devices()  # Ne devrait pas être appelé directement dans une route Flask, mais plutôt en tant que tâche asynchrone
    return redirect(url_for('get_client1'))

@app.route('/open_putty/<client>/<device_id>', methods=['GET'])
def open_putty(client, device_id):
    cur = mysql.connection.cursor(DictCursor)
    if client == 'client1':
        cur.execute("SELECT ip, username FROM client1 WHERE id = %s", (device_id,))
        device_info = cur.fetchone()
    elif client == 'client2':
        cur.execute("SELECT ip, username FROM client2 WHERE id = %s", (device_id,))
        device_info = cur.fetchone()
    cur.close()

    if device_info:
        username = device_info['username']
        ip = device_info['ip']
        # putty_path = r"C:\Users\user\Downloads\Putty1.exe"

        app_putty = Application().start(f"putty.exe -ssh {username}@{ip}")
        putty = app_putty.window(title_re=".*PuTTY")
        
        # Enter the password and press Enter
        putty.type_keys("password{ENTER}", with_spaces=True)
        time.sleep(1)
        putty.type_keys("ls{ENTER}", with_spaces=True)

        return jsonify({'message': 'PuTTY opened successfully'}), 200
    else:
        return jsonify({'message': 'Device not found'}), 404

def get_devices():
    url = "http://192.168.56.10/legacy/lab1.unl/topology"
    username = "admin"
    password = "eve"

    response = requests.get(url, auth=(username, password))
    if response.status_code == 200:
        devices = response.json()
        return devices
    else:
        return []

@app.route('/getcisco/<host>/<username>/<password>', methods=['POST'])
def get_cisco_info(host, username, password):
    device = {
        'device_type': 'autodetect',
        'host': host,
        'username': username,
        'password': password,
    }
    try:
        connection = ConnectHandler(**device)
        output = connection.send_command('show tech-support')
        output1 = connection.send_command('show running-config')

        info = {
            'equipement': '',
            'AdresseIP': host,
            'Fabricant': 'Unknown',
            'Fonction': 'Unknown',
            'Role': 'Unknown',
            'Version': 'Unknown',
            'SerialNumber': '',
            'EvaluationLicenseExpires': '',
            'Addressemac': '',

        }

        match_hostname = re.search(r'hostname\s+(\S+)', output, re.IGNORECASE)
        if match_hostname:
            info['equipement'] = match_hostname.group(1)

        match_version = re.search(r'version\s+(\S+)', output1, re.IGNORECASE)
        if match_version:
            info['Version'] = match_version.group(1)
        
        match_mac = re.search(r'Hardware is \S+, address is ([0-9A-Fa-f.]+)', output)
        if match_mac:
            info['Addressemac'] = match_mac.group(1).replace('.', ':')

        if 'firewall' in output.lower():
            info['Fonction'] = 'Sécurité'
            info['Role'] = 'Firewall'
        elif 'switch' in output.lower():
            info['Fonction'] = 'Réseau'
            info['Role'] = 'Switch'
        elif 'router' in output.lower():
            info['Fonction'] = 'Réseau'
            info['Role'] = 'Routeur'
        elif 'wifi' in output.lower():
            info['Fonction'] = 'Réseau'
            info['Role'] = 'Wifi'

        if 'cisco' in output.lower():
            info['Fabricant'] = 'Cisco'
        elif 'fortinet' in output.lower():
            info['Fabricant'] = 'Fortinet'
        elif 'aruba' in output.lower():
            info['Fabricant'] = 'Aruba'
        elif 'sophos' in output.lower():
            info['Fabricant'] = 'Sophos'

        connection.disconnect()
        
        return jsonify({"message": info})

    except Exception as e:
        print(f"Erreur lors de la connexion à {host}: {str(e)}")
        return jsonify({"message": f"Erreur lors de la connexion à {host}: {str(e)}"})
    

@app.route('/getfortigate/<host>/<username>/<password>', methods=['POST'])
def get_fortigate_info(host, username, password):
    """
    Récupère les informations d'un équipement FortiGate.
    """
    device = {
        'device_type': 'fortinet',
        'host': host,
        'username': username,
        'password': password,
    }
    connection = ConnectHandler(**device)
    output = connection.send_command('get system status')
    output1 = connection.send_command('get hardware nic port1')

    info = {
        'equipement': '',
        'AdresseIP': host,
        'Fabricant': 'Fortinet',
        'Fonction': 'Security',
        'Role': 'Firewall',
        'Version': 'Unknown',
        'SerialNumber': '',
        'EvaluationLicenseExpires': '',
        'Addressemac': '',
    }

    match_hostname = re.search(r'Hostname: (.+)', output)
    if match_hostname:
        info['equipement'] = match_hostname.group(1)

    match_version = re.search(r'(\bv\d+\.\d+\.\d+\b)', output)
    if match_version:
        info['Version'] = match_version.group(1)
    match_serial_number = re.search(r'Serial-Number: (.+)', output)
    if match_serial_number:
        info['SerialNumber'] = match_serial_number.group(1)

    match_license_expires = re.search(r'Evaluation License Expires: (.+)', output)
    if match_license_expires:
        info['EvaluationLicenseExpires'] = match_license_expires.group(1)

    match_mac = re.search(r'Hwaddr:\s+([0-9A-Fa-f:]+)', output1)
    if match_mac:
        info['Addressemac'] = match_mac.group(1)

    connection.disconnect()

    return jsonify({"message":info})


def ssh_connect(ip, username, password):
    """Se connecter à un équipement réseau via SSH et exécuter une commande."""
    print("##################ssh_connect(ip, username, password)")
    try:
        client = retrieve_config_terminal.SSHClient()
        client.set_missing_host_key_policy(retrieve_config_terminal.AutoAddPolicy())
        client.connect(ip, username=username, password=password)
        logging.debug("Connexion SSH réussie à %s", ip)

        stdin, stdout, stderr = client.exec_command("show tech-support")
        output = stdout.readlines()
        client.close()

        return output
    except retrieve_config_terminal.AuthenticationException:
        logging.error("Échec de l'authentification SSH à %s", ip)
    except retrieve_config_terminal.SSHException as e:
        logging.error("Erreur SSH lors de la connexion à %s : %s", ip, e)
    except Exception as e:
        logging.error("Erreur lors de la connexion à %s : %s", ip, e)
    return []

def process_output(output):
    """Traiter la sortie de la commande SSH pour extraire les données nécessaires."""
    print("################process_output(output)")
    data = []
    category, point_de_controle, description, impact, conformite, criticite, remediation = "", "", "", "", "", "", ""
    
    for line in output:
        if "Catégorie" in line:
            category = line.split(":")[1].strip()
        elif "Point de contrôle" in line:
            point_de_controle = line.split(":")[1].strip()
        elif "Description" in line:
            description = line.split(":")[1].strip()
        elif "Impact" in line:
            impact = line.split(":")[1].strip()
        elif "Conformité" in line:
            conformite = line.split(":")[1].strip()
        elif "Criticité" in line:
            criticite = line.split(":")[1].strip()
        elif "Remédiation" in line:
            remediation = line.split(":")[1].strip()
            data.append((category, point_de_controle, description, impact, conformite, criticite, remediation))
    
    # Ajout de points de contrôle spécifiques
    add_specific_checkpoints(data, output)
    
    return data

def add_specific_checkpoints(data, output):
    """Ajouter des points de contrôle spécifiques basés sur les lignes de sortie."""
    print("##################add_specific_checkpoints(data, output)")
    checkpoints = [
        ("Management Plane", "Activation du AAA", 
         "Cette commande active le système de contrôle d'accès AAA", 
         "L'activation du modèle AAA entraîne la désactivation immédiatement des méthodes d'accès précédentes afin de garantir la continuité des services et de minimiser les perturbations pour les utilisateurs.", 
         "conforme" if any("aaa new-model" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'aaa new-model'."),
         
        ("Management Plane", "Activation de l'authentification AAA", 
         "Cette commande configure l'authentification AAA à des fins de connexion", 
         "L'activation de l'authentification AAA entraîne immédiatement la désactivation des méthodes d'accès précédentes, garantissant une transition en douceur vers des méthodes plus sécurisées tout en minimisant les perturbations pour les utilisateurs.", 
         "conforme" if any("aaa authentication login" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'aaa authentication login {default | aaa_list_name} [passwd-expiry] [method1] [method2]'."),
        
        ("Management Plane", "Authentification de connexion pour la ligne console 0", 
         "Authentifie les utilisateurs qui accèdent au routeur ou au commutateur à l'aide du port de console série", 
         "L'activation de l'authentification AAA Cisco 'line login' est significativement perturbatrice car les méthodes d'accès précédentes sont immédiatement désactivées.", 
         "conforme" if any("login authentication" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'login authentication {default | aaa_list_name}' "),
        
        ("Management Plane", "Authentification de connexion pour le terminal de ligne", 
         "Authentifie les utilisateurs qui accèdent au routeur ou au commutateur à l'aide du port TTY.", 
         "L'activation de l'authentification de connexion pour la ligne TTY avec Cisco AAA entraîne une perturbation significative car les anciennes méthodes d'accès sont immédiatement désactivées.", 
         "conforme" if any("line tty" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'line tty {line-number} >> login authentication {default | aaa_list_name}'"),
        
        ("Management Plane", "Authentification de connexion pour la ligne vty", 
         "Authentifie les utilisateurs qui accèdent au routeur ou commutent à distance via le port VTY.", 
         "L'activation de l'authentification de connexion pour la ligne VTY avec Cisco AAA entraîne une perturbation significative car les anciennes méthodes d'accès sont immédiatement désactivées.", 
         "conforme" if any("line vty " in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'line vty {line-number} [<em>ending-line-number]>>login authentication {default | aaa_list_name}'"),
        
        ("Management Plane", "'aaa accounting' pour enregistrer toutes les commandes d'utilisation privilégiéeà l'aide de 'commandes 15'", 
         "Exécute la comptabilité de toutes les commandes au niveau de privilège spécifié.", 
         "L'activation de la « comptabilité AAA » pour les commandes privilégiées permet de suivre et de documenter l'activité des utilisateurs, renforçant ainsi la sécurité et facilitant la détection des anomalies et la conformité réglementaire.", 
         "conforme" if any("aaa accounting commands 15" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'aaa accounting commands 15' pour enregistrer toutes les commandes d'utilisation privilégiée."),
        

        ("Management Plane", "'aaa accounting connection'", 
         "Fournit des informations sur toutes les connexions sortantes établies à partir du serveur d'accès réseau.", 
         "Une surveillance régulière des enregistrements de connexion pour détecter les anomalies, résoudre les problèmes et assurer la conformité réglementaire.", 
         "conforme" if any("aaa accounting connection" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'aaa accounting connection'."),

        
        ("Management Plane", "'aaa accounting exec'", 
         "Exécute la comptabilité pour la session shell EXEC.", 
         "La surveillance des sessions de terminal, la détection des activités suspectes, et garantit la conformité réglementaire via des enregistrements précis des utilisateurs, heures de connexion et activités.", 
         "conforme" if any("aaa accounting exec" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'aaa accounting exec'."),

        ("Management Plane", "'aaa accounting network'", 
         "Exécute la comptabilité de toutes les demandes de service liées au réseau.", 
         "La création des enregistrements comptables et leur surveillance régulière permettent de détecter les exceptions, de résoudre les problèmes et de signaler les résultats.", 
         "conforme" if any("aaa accounting network " in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'aaa accounting network'."),
        
        ("Management Plane", "'aaa accounting system'", 
         "Effectue la comptabilisation de tous les événements au niveau du système non associés aux utilisateurs, tels que les rechargements.", 
         "Une surveillance régulière des enregistrements pour identifier les anomalies, résoudre les problèmes et rapporter les résultats de manière continue.", 
         "conforme" if any(" aaa accounting system " in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'aaa accounting system'."),

        ("Management Plane", "le « privilège 1 » pour les utilisateurs locaux", 
         "Définit le niveau de privilège de l'utilisateur.", 
         "réduire le risque d’accès non autorisé.", 
         "conforme" if any(" privilege 1 " in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'username <LOCAL_USERNAME> privilege 1' "),

        ("Management Plane", "transport input ssh' pour les connexions 'line vty'", 
         "Sélectionne le protocole Secure Shell (SSH).", 
         "réduire le risque d’accès non autorisé.", 
         "conforme" if any(" transport input ssh" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'transport input ssh' "),

        ("Management Plane", "'no exec' pour 'line aux 0'", 
         "La commande 'no exec' restreint une ligne aux connexions sortantes uniquement.", 
         "Réduit le risque d'accès non autorisé en désactivant le port 'aux' avec la commande 'no exec'.", 
         "conforme" if any(" no exec" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'line aux 0 '>>'no exec' "),

        ("Management Plane", "'access-list' à utiliser avec 'line vty'", 
         "Les listes d'accès contrôlent la transmission des paquets sur une interface, contrôlent l'accès à la ligne de terminal virtuel (VTY) et restreignent le contenu des mises à jour de routage.", 
         "Réduire le risque d'accès non autorisé en mettant en œuvre des listes d'accès pour toutes les lignes VTY.", 
         "conforme" if any(" ip access-list" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'access-list' "),

        ("Management Plane", "'access-class' for 'line vty'", 
         "Le paramètre « access-class » restreint les connexions entrantes et sortantes entre un VTY et les périphériques réseau associés aux adresses dans une liste d'accès.", 
         "Restreindre l'accès à distance aux seuls appareils autorisés à gérer le périphérique réduit ainsi le risque d'accès non autorisé.", 
         "conforme" if any("access-class" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'access-class' "),

        ("Management Plane", "'exec-timeout' sur une valeur inférieure ou égale à 10 minutes pour 'line aux 0' ", 
         "Si aucune entrée n'est détectée pendant l'intervalle, la fonction EXEC reprend la connexion en cours ou, s'il n'y a pas de connexion, met le terminal en état inactif et déconnecte la session entrante.", 
         "L'activation du « délai d'attente d'exécution » avec une durée appropriée de minutes ou de secondes empêche l'accès non autorisé aux sessions abandonnées.", 
         "conforme" if any("exec-timeout" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'line aux 0'>> 'exec-timeout <timeout>' "),

        ("Management Plane", "'exec-timeout' sur une valeur inférieure ou égale à 10 minutes pour la console de ligne 0'", 
         "La fonction 'exec' reprend la connexion en cours si aucune entrée n'est détectée pendant l'intervalle spécifié. Si aucune connexion n'est établie, elle ramène le terminal à l'état inactif et déconnecte la session entrante.", 
         "L'activation du « délai d'attente d'exécution » avec une durée appropriée de minutes ou de secondes empêche l'accès non autorisé aux sessions abandonnées.", 
         "conforme" if any("exec-timeout" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'line con 0' >> 'exec-timeout <timeout>' "),

        ("Management Plane", "''exec-timeout' inférieur ou égal à 10 minutes 'line tty'", 
         "La fonction 'exec' reprend la connexion en cours si aucune entrée n'est détectée pendant l'intervalle spécifié. Si aucune connexion n'est établie, elle ramène le terminal à l'état inactif et déconnecte la session entrante.", 
         "L'activation du « délai d'attente d'exécution » avec une durée appropriée de minutes ou de secondes empêche l'accès non autorisé aux sessions abandonnées.", 
         "conforme" if any("line tty 0 | begin Timeout" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'line tty <numéro> | begin Timeout ' >> ' exec-timeout <timeout>' "),

        ("Management Plane", "'exec-timeout' sur une valeur inférieure ou égale à 10 minutes 'line vty'", 
         "La fonction 'exec' reprend la connexion en cours si aucune entrée n'est détectée pendant l'intervalle spécifié. Si aucune connexion n'est établie, elle ramène le terminal à l'état inactif et déconnecte la session entrante.", 
         "L'activation du « délai d'attente d'exécution » avec une durée appropriée de minutes ou de secondes empêche l'accès non autorisé aux sessions abandonnées.", 
         "conforme" if any(" line vty 0 | begin Timeout" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'line vty {line_number}' >> 'exec-timeout <<span>timeout_in_minutes> ' "),

        ("Management Plane", "'transport input none' for 'line aux 0'", 
         "autoriser uniquement une connexion sortante sur une ligne", 
         "La désactivation de tous les protocoles entrants sur les ports auxiliaires vise à prévenir tout accès non autorisé.", 
         "conforme" if any(" transport input none " in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'line aux 0' >> ' transport input none ' "),

        ("Management Plane", "'banner-text' for 'banner exec'", 
         "Cette commande spécifie un message à afficher lorsqu'un processus EXEC est créé (une ligne est activée ou une connexion entrante est établie vers un vty).", 
         "Fournir des avis juridiques et des avertissements appropriés aux personnes accédant à leurs réseaux en utilisant un « texte de bannière » pour la commande de bannière d'exécution.", 
         "conforme" if any("banner exec" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'banner exec <caractère_délimitateur><votre_message><caractère_délimitateur>' "),

        ("Management Plane", "'banner-text' for 'banner login'", 
         "Cette commande spécifie un message à afficher lorsqu'un processus EXEC est créé (une ligne est activée ou une connexion entrante est établie vers un vty).", 
         "Cette commande spécifie un message à afficher lorsqu'un utilisateur tente de se connecter au routeur ou switch ", 
         "conforme" if any("banner login" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'banner login <character>'>> '<banner-text> <character>' "),

        ("Management Plane", "'banner-text' for 'banner motd'", 
         "La bannière MOTD s'affiche sur tous les terminaux connectés et permet d'envoyer des messages impactant tous les utilisateurs, tels que des notifications concernant les arrêts imminents du système.", 
         "fournir des avis juridiques et des avertissements appropriés aux personnes accédant à leurs réseaux, les administrateurs utilisent un 'texte de bannière' en configurant la commande 'banner motd'. ", 
         "conforme" if any("banner motd" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'banner motd <character>'>> '<banner-text> <character>' "),

        ("Management Plane", " 'password' for 'enable secret' ", 
         "fournir une couche de sécurité supplémentaire pour le mot de passe d'activation en le stockant de manière cryptographique irréversible.", 
         "protéger l'accès au mode privilégié",  
         "conforme" if any("enable secret" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande ' enable secret {ENABLE_SECRET_PASSWORD} ' "),

        ("Management Plane", "'service password-encryption'", 
         "La forme cryptée des mots de passe s'affiche lorsqu'une commande more system:running-config est saisie.", 
         "Réduire le risque que des utilisateurs non autorisés puissent accéder aux mots de passe en texte clair dans les fichiers de configuration Cisco IOS.",  
         "conforme" if any("service password-encryption" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande ' service password-encryption ' "),

        ("Management Plane", "'username secret' pour tous les utilisateurs locaux", 
         "Utilisez la commande 'username' avec l'option 'secret' pour configurer un nom d'utilisateur et un mot de passe utilisateur chiffré en MD5.", 
         "Réduire le risque que des utilisateurs non autorisés accèdent aux appareils Cisco IOS.",  
         "conforme" if any("username" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande ' username {{em}LOCAL_USERNAME{/em}} secret {{em}LOCAL_PASSWORD{/em}}' "),

        ("Management Plane", "'no snmp-server'pour désactiver SNMP lorsqu'il n'est pas utilisé", 
         "S'il n'est pas utilisé, désactivez le protocole SNMP, l'accès en lecture et en écriture.", 
         "Le SNMP est un protocole utilisé pour surveiller et gérer les périphériques réseau à distance. Cependant, s'il n'est pas utilisé, il est recommandé de le désactiver, y compris les accès en lecture et en écriture.",  
         "conforme" if any("no snmp-server" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande ' no snmp-server ' "),

        ("Management Plane", "Désactiver 'private' pour 'snmp-server community'", 
         "Une chaîne de communauté SNMP permet un accès en lecture seule à tous les objets.", 
         "La désactivation de la chaîne de communauté par défaut 'private' réduit le risque d'accès non autorisé au périphérique via SNMP",  
         "conforme" if any("no snmp-server community private" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande ' no snmp-server community private ' "),

        ("Management Plane", "Désactiver 'public' pour 'snmp-server community'", 
         "Une chaîne de communauté SNMP permet un accès en lecture seule à tous les objets.", 
         "La désactivation de la chaîne de communauté par défaut 'public' réduit le risque d'accès non autorisé au périphérique via SNMP",  
         "conforme" if any("no snmp-server community public" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande ' no snmp-server community public' "),

        ("Management Plane", "Ne définissez pas 'RW' pour aucune 'snmp-server community'", 
         "Les stations de gestion autorisées ayant cette chaîne de communauté pourront à la fois récupérer et modifier les objets MIB.", 
         "Pour réduire le risque d'accès non autorisé, il est recommandé de désactiver l'accès en écriture SNMP pour la communauté du serveur SNMP.",  
         "conforme" if all("RW" not in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande ' no snmp-server community {write_community_string} ' "),

        ("Management Plane", "Configuration de l'ACL pour chaque 'snmp-server community'", 
         "Cette fonctionnalité spécifie une liste d'adresses IP autorisées à utiliser la chaîne de communauté pour accéder à l'agent SNMP.", 
         "Pour réduire le risque d'accès non autorisé, activez les listes de contrôle d'accès pour toutes les communautés SNMP des serveurs et restreignez l'accès aux zones de gestion de confiance appropriées.",  
         "conforme" if all("snmp-server community" in line and "ro" in line and "snmp_access-list_number" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'snmp-server community <community_string> ro <snmp_access-list_number> '"),

        ("Management Plane", "Configuration de 'snmp-server enable traps snmp' ", 
         "Les notifications SNMP peuvent être envoyées sous forme d'interruptions aux systèmes de gestion autorisés.", 
         "limiter l'envoi de messages SNMP uniquement aux systèmes explicitement nommés afin de réduire les accès non autorisés.",  
         "conforme" if any("snmp-server enable traps snmp" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'snmp-server enable traps snmp authentication linkup linkdown coldstart' "),

        ("Management Plane", "Définir 'priv' pour chaque 'snmp-server group' à l'aide de SNMPv3", 
         "Spécifie l'authentification d'un paquet avec cryptage lors de l'utilisation de SNMPv3.", 
         "Réduire considérablement les risques d'accès non autorisé en utilisant le paramètre « snmp-server group v3 priv » pour chiffrer les messages en transit.",  
         "conforme" if any("snmp-server group" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'snmp-server group {<em>group_name</em>} v3 priv' "),

        ("Management Plane", "Exiger 'aes 128' au minimum pour 'snmp-server user' lors de l'utilisation SNMPv3", 
         "Spécifiez l'utilisation d'un algorithme AES d'au moins 128 bits pour le chiffrement lors de l'utilisation de SNMPv3.", 
         "Réduire considérablement les risques d'accès non autorisé en utilisant le paramètre 'snmp-server user' avec des protocoles d'authentification et de confidentialité appropriés pour chiffrer les messages en transit.",  
         "conforme" if any("snmp-server user" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'snmp-server user {user_name} {group_name} v3 auth sha {auth_password} priv aes 128 {priv_password} {acl_name_or_number} ' "),

        ("Control Plane", "Définir la version 2 pour la 'ip ssh version'", 
         "La version de SSH à exécuter.", 
         "Réduire le risque d'accès non autorisé.",  
         "conforme" if any("ip ssh version 2" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ip ssh version 2 ' "),

        ("Control Plane", "Configuration de'no cdp run' ", 
         "Désactivez le service Cisco Discovery Protocol (CDP). ", 
         "Réduire les risques d'accès non autorisé, limiter l'utilisation des protocoles réseau et exiger la désactivation explicite des protocoles non sécurisés ou non nécessaires.",  
         "conforme" if any("no cdp run" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'no cdp run ' "),

        ("Control Plane", "Configuration de 'no ip bootp server'  ", 
         "Désactivez le service Bootstrap Protocol (BOOTP). ", 
         "Réduire les risques d'accès non autorisé, limiter l'utilisation des protocoles réseau et exiger la désactivation explicite des protocoles non sécurisés ou inutiles tels que le « serveur ip bootp ».",  
         "conforme" if any("ip dhcp bootp ignore" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ip dhcp bootp ignore ' "),

        ("Control Plane", "Configuration de'no service dhcp'  ", 
         "Désactivez les fonctionnalités du serveur DHCP et de l'agent de relais ", 
         "Réduire les risques d'accès non autorisé, limiter l'utilisation des protocoles réseau et exiger la désactivation explicite des protocoles non sécurisés ou  inutiles tels que le protocole DHCP.",  
         "conforme" if any("no service dhcp" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'no service dhcp' "),

        ("Control Plane", "Configuration de'no ip identd'  ", 
         "Désactivez le serveur d'identification (identd). ", 
         "Réduire les risques d'accès non autorisé, limiter l'utilisation des protocoles réseau et exiger la désactivation explicite des protocoles non sécurisés ou inutiles tels que le protocole d'identification (identd).",  
         "conforme" if any("no ip identd" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'no ip identd ' "),

        ("Control Plane", "Configuration de 'service tcp-keepalives-in'  ", 
         "Générez des paquets keepalive sur les connexions réseau entrantes inactives. ", 
         "Réduire le risque d'accès non autorisé, limiter la durée d'autorisation des sessions terminées et appliquer cette politique via l'utilisation de la commande « tcp-keepalives-in ».",  
         "conforme" if any("service tcp-keepalives-in" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'service tcp-keepalives-in ' "),

        ("Control Plane", "Configuration de'service tcp-keepalives-out' ", 
         "Générez des paquets keepalive sur les connexions réseau sortantes inactives. ", 
         "Réduire le risque d'accès non autorisé, limiter la durée d'autorisation des sessions terminées et appliquer cette politique via l'utilisation de la commande « tcp-keepalives-out ».",  
         "conforme" if any("service tcp-keepalives-out" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'service tcp-keepalives-out' "),

        ("Control Plane", "Configuration de'no service pad' ", 
         "Désactivez le service d’assemblage/désassemblage de paquets (PAD) .", 
         "Réduire le risque d'accès non autorisé, les organisations doivent mettre en œuvre une politique de sécurité restreignant les services inutiles tels que le service « PAD ».",  
         "conforme" if any("no service pad" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'no service pad' "),
        
        ("Control Plane", "Configuration de'logging on' ", 
         "Activer la journalisation des messages système.", 
         "la surveillance des risques technologiques .",  
         "conforme" if any("logging enable" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'logging enable' "),

        ("Control Plane", "Configuration de'buffer size' for 'logging buffered' ", 
         "Activez la journalisation des messages système dans un tampon local.", 
         "gérer les risques technologiques ",  
         "conforme" if any("logging buffered" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'logging buffered [<em>log_buffer_size</em>]' "),

        ("Control Plane", "Configuration de'logging console critical'  ", 
         "Vérifiez que la journalisation sur la console de l'appareil est activée et limitée à un niveau de gravité rationnel .", 
         "gérer les risques technologiques , capturer les messages de gravité ",  
         "conforme" if any("logging console critical" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'logging console critical' "),

        ("Control Plane", "Définir l'adresse IP pour 'logging host'  ", 
         "Enregistrez les messages système et la sortie de débogage sur un hôte distant.", 
         "gérer les risques technologiques , définir l'adresse IP de l'hôte de journalisation et applique le processus de journalisation. ",  
         "conforme" if any("logging host" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'logging host {syslog_server}' "),

        ("Control Plane", "Configuration de'logging trap informational'  ", 
         "Limitez les messages consignés sur les serveurs Syslog en fonction du niveau de gravité des informations.", 
         "gérer les risques technologiques , La commande 'logging trap' définit la gravité des messages et applique le processus de journalisation. ",  
         "conforme" if any("logging trap informational" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'logging trap informational' "),

        ("Control Plane", "Configuration de'service timestamps debug datetime'   ", 
         "Limitez les messages consignés sur les serveurs Syslog en fonction du niveau de gravité des informations.", 
         "gérer les risques technologiques , et l'établissement d'une chronologie des événements est essentiel. La commande « service timestamps » définit la date et l'heure des entrées envoyées à l'hôte de journalisation et applique le processus de journalisation. ",  
         "conforme" if any("service timestamps" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'service timestamps debug datetime' "),

        ("Control Plane", "Configuration de'logging source interface'  ", 
         "Spécifiez l'adresse IPv4 ou IPv6 source des paquets de journalisation système.", 
         "gérer les risques technologiques , La commande « logging source interface loopback » définit une adresse IP cohérente pour envoyer des messages à l'hôte de journalisation et applique le processus de journalisation. ",  
         "conforme" if any("logging source interface" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'logging source-interface loopback' "),

        ("Control Plane", "Configuration de'ntp authenticate'", 
         "Activez l'authentification NTP.", 
         "réduire les risques liés à une synchronisation incorrecte ou malveillante.  ",  
         "conforme" if any("ntp authenticate" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ntp authenticate' "),

        ("Control Plane", "Configuration de'ntp authentication-key'", 
         "Définissez une clé d'authentification pour NTP.", 
         "offre une protection contre les attaques basées sur le temps.  ",  
         "conforme" if any("ntp authentication-key" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ntp authentication-key {ntp_key_id} md5 {ntp_key_hash}' "),

        ("Control Plane", "Configuration de'ntp trusted-key'", 
         "Assurez-vous d'authentifier l'identité d'un système sur lequel le NTP sera synchronisé.", 
         "offre une protection contre les attaques basées sur le temps.  ",  
         "conforme" if any("ntp trusted-key" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ntp trusted-key {ntp_key_id}' "),

        ("Control Plane", "Configuration de 'key' pour chaque 'ntp server'", 
         "Spécifie la clé d'authentification pour NTP.", 
         "Protection contre les attaques basées sur le temps .  ",  
         "conforme" if any("ntp server" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ntp server {<em>ntp-server_ip_address</em>}{key  <em>ntp_key_id</em>}' "),
        
        ("Control Plane", "Configuration de'ip address' for 'ntp server'", 
         "autoriser le système à synchroniser l'horloge du logiciel système avec le serveur NTP spécifié.", 
         "Protection contre les attaques basées sur le temps .  ",  
         "conforme" if any("ntp server" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ntp server {ntp-server_ip_address}' "),

        ("Control Plane", "Créer un seul 'interface loopback'", 
         "Configurer une seule interface de bouclage.", 
         "Sécurisation des Services Critiques.",  
         "conforme" if any("interface loopback" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'interface loopback {number} ip address {loopback_ip_address}' "),

        ("Control Plane", "Configuration de l'interface source AAA",
         "Force AAA à utiliser l'adresse IP d'une interface spécifiée pour tous les paquets AAA sortants.", 
         "Une surveillance efficace des dispositifs réseau, la simplicité de la gestion des politiques, la traçabilité, et la cohérence des configurations AAA dans un réseau.",  
         "conforme" if any("tacacs source" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ip radius source-interface loopback' "),

        ("Control Plane", "Configuration de'ntp source' sur l'interface de bouclage", 
         "Utilisez une adresse source particulière dans les paquets NTP.", 
         "la simplicité de la configuration et la sécurité de la synchronisation du temps dans un réseau.",  
         "conforme" if any("ntp source loopback" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande'ntp source loopback {<em>loopback_interface_number}</em>' "),

         ("Control Plane", "Configuration de'ip tftp source-interface' sur l'interface de bouclage", 
         "Spécifiez l'adresse IP d'une interface en tant qu'adresse source pour les connexions TFTP.", 
         "la simplicité de la configuration et la gestion de la connectivité et la sécurité des transferts TFTP dans un réseau.",  
         "conforme" if any("tftp source-interface loopback" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ip tftp source-interface loopback {<em>loopback_interface_number}</em>' "),

        ("Data Plane", "Configuration de'no ip source-route'", 
         "Désactive la gestion des datagrammes IP avec des options d'en-tête de routage source.", 
         "prévenir les attaques de type IP spoofing, à améliorer la gestion du trafic réseau et à simplifier la configuration des périphériques réseau.",  
         "conforme" if any("no ip source-route" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'no ip source-route' "),
         
        ("Data Plane", "Configuration de 'no ip source-route'", 
         "Désactiver le proxy ARP sur toutes les interfaces.", 
         "Prévention des attaques ARP Spoofing , Désactivation des services inutiles ",  
         "conforme" if any("no ip proxy-arp" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'no ip proxy-arp' pour Désactiver le proxy ARP sur toutes les interfaces"),

        ("Data Plane", "Configuration de 'no interface tunnel'", 
         "Vérifiez qu'aucune interface de tunnel n'est définie.", 
         "Réduire les surfaces d'attaque potentielles.les interfaces de tunnel, bien qu'utiles dans certains scénarios, peuvent représenter une vulnérabilité si elles ne sont pas correctement sécurisées ou si elles ne sont plus nécessaires. "
         "conforme" if any("no interface tunnel" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'no interface tunnel {instance}'"),

        ("Data Plane", "Configuration de 'ip verify unicast source reachable-via'", 
         "Examine les paquets entrants pour déterminer si l'adresse source est dans la table de routage (FIB) et autorise le paquet uniquement si la source est accessible via l'interface sur laquelle le paquet a été reçu .", 
         "protéger la confidentialité, l'intégrité et la disponibilité des appareils réseau. La fonctionnalité 'unicast Reverse-Path Forwarding' (uRPF) utilise dynamiquement la table de routage du routeur pour accepter ou rejeter les paquets lorsqu'ils arrivent sur une interface.. "
         "conforme" if any("ip verify unicast source reachable-via rx" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ip verify unicast source reachable-via rx'"),

        ("Data Plane", "Configuration de 'ip access-list extended'", 
         "Définir les conditions d'accès refusées ou autorisées à l'aide des commandes deny et permit.", 
         "L'ajout de 'ip access-list' autorisant et refusant explicitement les réseaux internes et externes applique ces politiques. "
         "conforme" if any("ip access-list extended" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ip access-list extended {<span><em>name | number</em>}'"),

        ("Data Plane", "Configuration de 'ip access-group' entrant sur l'interface externe", 
         "Définir les conditions d'accès refusées ou autorisées à l'aide des commandes deny et permit.", 
         "permettent et refusent explicitement l'accès en fonction des listes d'accès. L'utilisation de la commande 'ip access-group' applique ces politiques en identifiant explicitement les groupes autorisés."
         "conforme" if any("ip access-group " in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'ip access-group {name | number} in'"),

        ("Data Plane", "Configuration de 'key chain' ", 
         "Définir une chaîne de clés d'authentification pour activer l'authentification des protocoles de routage. Remarque : Seuls les protocoles DRP Agent, EIGRP, et RIPv2 utilisent des chaînes de clés.", 
         "Authentification des protocoles de routage , Protection contre les attaques."
         "conforme" if any("key chain" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'key chain {<em>key-chain_name</em>}'"),

        ("Data Plane", "Configuration de 'key' ", 
         "Configurer une clé d'authentification sur une chaîne de clés.", 
         "Authentification des protocoles de routage , Protection contre les attaques."
         "conforme" if any("key chain" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'key {<em>key-number</em>}'"),
        
        ("Data Plane", "Configuration de 'key-string' ", 
         "Configurer la chaîne d'authentification pour une clé.", 
         "Authentification des protocoles de routage , Protection contre les attaques."
         "conforme" if any("key chain" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'key-string <<em>key-string</em>>'"),

        ("Data Plane", "Configuration de 'address-family ipv4 autonome-system' ", 
         "Configurer la famille d'adresses EIGRP.", 
         "segmenter le trafic, soutenir différents protocoles de routage, optimiser le routage, améliorer la redondance et la tolérance aux pannes, renforcer la sécurité et le contrôle d'accès, faciliter la gestion des adresses IP. "
         "conforme" if any("address-family ipv4 autonomous-system" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'router eigrp <<em>virtual-instance-name</em>>','address-family ipv4 autonomous-system {<em>eigrp_as-number</em>} '"),
        
        ("Data Plane", "Définition de 'af-interface default' ", 
         "Définit les paramètres par défaut à appliquer aux interfaces EIGRP appartenant à une famille d'adresses.", 
         "segmenter le trafic, soutenir différents protocoles de routage, optimiser le routage, améliorer la redondance et la tolérance aux pannes, renforcer la sécurité et le contrôle d'accès, faciliter la gestion des adresses IP."
         "conforme" if any("af-interface default" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'router eigrp <<em>virtual-instance-name</em>>','address-family ipv4 autonomous-system {<em>eigrp_as-number</em>} ', 'af-interface default'"),

        ("Data Plane", "Configuration de 'authentification key-chain'", 
         "Configurez la chaîne de clés EIGRP pour la famille d'adresses.", 
         "segmenter le trafic, soutenir différents protocoles de routage, optimiser le routage, améliorer la redondance et la tolérance aux pannes, renforcer la sécurité et le contrôle d'accès, faciliter la gestion des adresses IP."
         "conforme" if any("authentication mode md5" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'router eigrp <<em>virtual-instance-name</em>>','address-family ipv4 autonomous-system {<em>eigrp_as-number</em>} ', 'af-interface default',' authentication mode md5'"),

        ("Data Plane", "Configuration de 'authentification message-digest' pour la zone OSPF ", 
         "Activer l'authentification MD5 pour OSPF.", 
         "prévenir les attaques d'usurpation, les attaques de l'homme du milieu et les accès non autorisés au réseau OSPF."
         "conforme" if any("authentication message-digest" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'router ospf <<em>ospf_process-id</em>>','area <<em>ospf_area-id</em>> authentication message-digest '"),
        
        ("Data Plane", "Configuration de 'authentification message-digest' pour la zone OSPF ", 
         "Activer l'authentification Message Digest 5 (MD5) pour OSPF.", 
         "prévenir les attaques d'usurpation, les attaques de l'homme du milieu et les accès non autorisés au réseau OSPF."
         "conforme" if any("ip ospf message-digest-key md5" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'router ospf <<em>ospf_process-id</em>>','ip ospf message-digest-key {<em>ospf_md5_key-id</em>} md5 {<em>ospf_md5_key</em>}'"),

        ("Data Plane", "Définition de 'key chain' ", 
         "définir une chaîne de clés d'authentification pour activer l'authentification des protocoles de routage RIPv2.", 
         "Authentifier les communications de routage pour les protocoles RIPv2 , gérer efficacement les clés cryptographiques."
         "conforme" if any("key chain" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'key chain {rip_key-chain_name}'"),

        ("Data Plane", "Configuration de 'key' ", 
         "Configurer une clé d'authentification sur une chaîne de clés.", 
         "Authentifier les communications de routage pour les protocoles RIPv2 , gérer efficacement les clés cryptographiques."
         "conforme" if any("key " in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'key {<em>key-number</em>}'"),
        
        ("Data Plane", "Configuration de 'key-string'", 
         "Configurer une clé d'authentification sur une chaîne de clés.", 
         "Authentifier les communications de routage pour les protocoles RIPv2 , gérer efficacement les clés cryptographiques."
         "conforme" if any("key-string" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'key-string <<em>key-string</em>>'"),

        ("Data Plane", "Configuration de 'IP RIP Authentication Key-Chain'", 
         "Activer l'authentification pour les paquets du protocole d'information de routage (RIP) Version 2 et spécifiquement l'ensemble de clés pouvant être utilisées sur une interface.", 
         "la protection contre les attaques, limite l'accès aux mises à jour de routage, améliore la fiabilité des communications RIP ."
         "conforme" if any("IP RIP Authentication Key-Chain" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'interface {<em>interface_name</em>}','ip rip authentication key-chain {<em>rip_key-chain_name</em>}'"),

        ("Data Plane", "Configuration de 'IP RIP Authentication Mode' en 'md5' ", 
         "Configurer l'interface avec la chaîne de clés RIPv2.", 
         "crypter les informations de routage RIP et garantir  l'authenticité des mises à jour RIP."
         "conforme" if any("ip rip authentication mode md5" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'interface <<em>interface_name</em>>','ip rip authentication mode md5'"),
        
        ("Data Plane", "Configuration de 'mot de passe voisin' ", 
         "Activer l'authentification message digest5 (MD5) sur une connexion TCP entre deux paires BGP.", 
         "L'utilisation du 'neighbor password' pour BGP renforce ces politiques en restreignant le type d'authentification entre les appareils réseau."
         "conforme" if any("router bgp" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'router bgp <<em>bgp_as-number</em>>','neighbor <<em>bgp_neighbor-ip</em> | <em>peer-group-name</em>> password <<em>password</em>> '"),  

    ]

    
    data.extend(checkpoints)

def generate_drawings(data):
    """Générer des graphiques basés sur les données traitées."""
    print("##################generate_drawings(data)")
    # Use a non-interactive backend
    plt.switch_backend('Agg')
    
    # Graphique à barres
    categories = [d[0] for d in data]
    errors = [d.count("Non conforme") for d in data]
    plt.figure(figsize=(8, 6))
    plt.bar(categories, errors)
    plt.title('nombre des erreurs par catégorie')
    plt.xlabel('categories')
    plt.ylabel("nombre d erreurs")
    plt.savefig("dessin_rapport_audit_switch.png")
    plt.close()  # Close the figure
    

    # Graphique circulaire
    labels = list(set(categories))
    sizes = [categories.count(label) for label in labels]
    plt.figure(figsize=(8, 6))
    plt.pie(sizes, labels=labels, autopct='%1.1%f%%')
    plt.axis('equal')
    plt.title('Répartition des erreurs par catégorie')
    plt.savefig("dessin_rapport_audit_switch.png")
    plt.close()  # Close the figure

def create_excel(data):
    """Créer un fichier Excel contenant les données du rapport d'audit."""
    print("####################create_excel(data)")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"rapport_audit_Switch_Cisco_{timestamp}.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rapport Audit"

    # En-têtes de colonne
    headers = ["Catégorie", "Point de contrôle", "Description", "Impact", "Conformité", "Criticité", "Remédiation"]
    sheet.append(headers)

    # Largeurs de colonne
    column_widths = [20, 20, 50, 50, 15, 15, 50]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Style de l'en-tête
    header_fill = PatternFill(start_color="EC407A", end_color="EC407A", fill_type="solid")
    header_font = Font(color="060605", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Écrire les données dans la feuille de calcul
    for row in data:
        sheet.append(row)

    # Parcourir chaque cellule pour trouver "Management Plane" et la colorer en bleu
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Management Plane" in str(cell.value):
                cell.fill = PatternFill(start_color="2F9AF8", end_color="2F9AF8", fill_type="solid")  # Bleu
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")  # Couleur de la police en blanc pour une meilleure visibilité


    # Parcourir chaque cellule pour trouver "Control Plane" et 
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Control Plane" in str(cell.value):
                cell.fill = PatternFill(start_color="f97434", end_color="f97434", fill_type="solid")  
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")  # Couleur de la police en blanc pour une meilleure visibilité



    # Parcourir chaque cellule pour trouver "Data Plane" et 
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Data Plane" in str(cell.value):
                cell.fill = PatternFill(start_color="fb6fe5", end_color="fb6fe5", fill_type="solid")  
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")  # Couleur de la police en blanc pour une meilleure visibilité
    
     # Styles conditionnels pour la conformité
    for row in sheet.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column == 5:  # Colonne de conformité
                if cell.value.lower() == "conforme":
                   cell.fill = PatternFill(start_color="08A91E", end_color="08A91E", fill_type="solid")  # Vert 
                elif cell.value.lower() == "non conforme":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rouge
                cell.font = Font(color="FFFFFF")  # Texte en blanc pour toutes les cellules colorées
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



    # Styles conditionnels pour la colonne Criticité 
    for row in sheet.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column == 6:  # Colonne 6
                if "Élevée" in str(cell.value):
                    cell.font = Font(color="FF0000")  # Rouge
                elif "Moyenne" in str(cell.value):
                    cell.font = Font(color="FFA500")  # Orange
                elif "Faible" in str(cell.value):
                    cell.font = Font(color="00FF00")  # Vert
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 

    # Ajuster la hauteur des lignes pour s'adapter au contenu
    adjust_row_height(sheet)

    workbook.save(filename=excel_filename)

def adjust_row_height(sheet):
    """Ajuster automatiquement la hauteur des lignes pour s'adapter au contenu."""
    print("##################adjust_row_height(sheet)")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_height = 0
        for cell in row:
            if cell.value:
                cell_value = str(cell.value)
                lines = cell_value.count('\n') + 1
                max_line_length = max(len(line) for line in cell_value.split('\n'))
                estimated_height = (lines + max_line_length // sheet.column_dimensions[cell.column_letter].width) * 15
                max_height = max(max_height, estimated_height)
        sheet.row_dimensions[row[0].row].height = max_height

@app.route('/hardeningRS/<ip>/<username>/<password>', methods=['POST'])
def hardenin(ip, username, password):
    while True:
        output = ssh_connect(ip, username, password)
        if output:
            data = process_output(output)
            create_excel(data)
            generate_drawings(data)
            return jsonify({"message": "Success", "data": data})
        else:
            print("Échec de connexion SSH. Veuillez réessayer.")
            return jsonify({"message": "SSH connection failed. Please try again."}), 500

def ssh_connectF(ip, username, password):
    """Se connecter à un équipement réseau via SSH et exécuter une commande."""
    print("##################ssh_connectF(ip, username, password)")
    try:
        client = retrieve_config_terminal.SSHClient()
        client.set_missing_host_key_policy(retrieve_config_terminal.AutoAddPolicy())
        client.connect(ip, username=username, password=password)
        logging.debug("Connexion SSH réussie à %s", ip)

        stdin, stdout, stderr = client.exec_command("show full-configuration")
        output = stdout.readlines()
        client.close()

        return output
    except retrieve_config_terminal.AuthenticationException:
        logging.error("Échec de l'authentification SSH à %s", ip)
    except retrieve_config_terminal.SSHException as e:
        logging.error("Erreur SSH lors de la connexion à %s : %s", ip, e)
    except Exception as e:
        logging.error("Erreur lors de la connexion à %s : %s", ip, e)
    return []

def process_outputF(output):
    """Traiter la sortie de la commande SSH pour extraire les données nécessaires."""
    print("######################process_outputF(output)")
    data = []
    category, point_de_controle, description, impact, conformite, criticite, remediation = "", "", "", "", "", "", ""
    
    for line in output:
        if "Catégorie" in line:
            split_line = line.split(":")
            if len(split_line) > 1:
                category = split_line[1].strip()
        elif "Point de contrôle" in line:
            split_line = line.split(":")
            if len(split_line) > 1:
                point_de_controle = split_line[1].strip()
        elif "Description" in line:
            split_line = line.split(":")
            if len(split_line) > 1:
                description = split_line[1].strip()
        elif "Impact" in line:
            split_line = line.split(":")
            if len(split_line) > 1:
                impact = split_line[1].strip()
        elif "Conformité" in line:
            split_line = line.split(":")
            if len(split_line) > 1:
                conformite = split_line[1].strip()
        elif "Criticité" in line:
            split_line = line.split(":")
            if len(split_line) > 1:
                criticite = split_line[1].strip()
        elif "Remédiation" in line:
            split_line = line.split(":")
            if len(split_line) > 1:
                remediation = split_line[1].strip()
                data.append((category, point_de_controle, description, impact, conformite, criticite, remediation))
    
    # Ajout de points de contrôle spécifiques
    add_specific_checkpointsF(data, output)
    
    return data


def add_specific_checkpointsF(data, output):
    """Ajouter des points de contrôle spécifiques basés sur les lignes de sortie."""
    print("#####################add_specific_checkpointsF(data, output)")

    checkpoints = [
        ("Network Settings", "le serveur DNS", 
         "traduire les noms d'hôtes en adresses IP.", 
         "éviter les attaques de type man-in-the-middle.", 
         "conforme" if any("config system dns\\s+set primary 8\\.8\\.8\\.8\\s+set secondary 8\\.8\\.4\\.4" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system dns ' \n 'set primary 8.8.8.8 ' \n 'set secondary 8.8.4.4'."),

        ("Network Settings", "le trafic intra-zone", 
         "pour garantir que seul le trafic spécifique et autorisé circule entre les réseaux de la même zone.", 
         "prévenir les attaques et les menaces internes, améliorer la visibilité et la gestion du trafic.", 
         "conforme" if any("config system zone\\s+edit DMZ\\s+set intrazone deny" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system zone ' \n 'set edit DMZ ' \n 'set intrazone deny'."),

        ("Network Settings", "les services liés à la gestion sur le port WAN", 
         "Activer tout service lié à la gestion sur l'interface WAN présente un risque élevé. Les services liés à la gestion tels que HTTPS, HTTP, ping, SSH, SNMP et Radius devraient être désactivés sur WAN..", 
         "Réduction de la surface d'attaque , Prévention des attaques de déni de service , Protection des informations sensibles.", 
         "conforme" if any("unselect allowaccess ping https ssh snmp http radius-acct " in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'unselect allowaccess ping https ssh snmp http radius-acct'."),

        ("System Settings", "la bannière de pré-connexion.", 
         "interdire l'accès non autorisé, fournir un avis de journalisation ou de surveillance.", 
         "la protection contre les accès non autorisés et l'amélioration de la traçabilité des activités des utilisateurs. ", 
         "conforme" if any(" set pre-login-banner enable" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande ' set pre-login-banner enable'."),

        ("System Settings", "la bannière de  post-connexion .", 
         "Configure la bannière après que les utilisateurs se sont connectés avec succès. ", 
         "la réduction des incidents de sécurité et la traçabilité des activités des utilisateurs ", 
         "conforme" if any("set post-login-banner enable" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande ' set post-login-banner enable'."),
        
        ("System Settings", "le fuseau horaire local .", 
         "Configure les informations de fuseau horaire local afin que l'heure affichée par le périphérique soit plus pertinente pour ceux qui le consultent.", 
         "la synchronisation temporelle, la coordination des opérations, la précision des rapports, la conformité réglementaire . ", 
         "conforme" if any("set timezone" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande ' set timezone ' ."),

        ("System Settings", "hostname", 
         "Modifie le nom d'hôte par défaut de l'appareil.", 
         "Un nom d'hôte configuré permet une identification claire des appareils, facilitant ainsi la gestion des actifs, la corrélation des logs et les déploiements de certificats.", 
         "conforme" if any("set hostname " in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'set hostname \"New_FGT1\"'."),

        ("System Settings", "Désactiver l'installation du Firmware et de la configuration USB", 
         "Désactive la fonctionnalité d'installation automatique du port USB pour la configuration et le firmware.", 
         "Réduire le risque d'injection de configurations ou de firmwares non autorisés via le port USB renforce la sécurité de l'appareil.", 
         "conforme" if any("config system auto-install" in line and "set auto-install-config disable" in line and "set auto-install-image disable" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system auto-install' \n 'set auto-install-config disable ' \n 'set auto-install-image disable '."),
               
        ("System Settings", "Désactiver les clés statiques pour TLS ", 
         "Désactive la prise en charge des clés statiques sur les sessions TLS se terminant sur le FortiGate.", 
         "Protection contre les attaques de type Man-in-the-Middle , Prévention des attaques par force brute .", 
         "conforme" if any("set ssl-static-key-ciphers disable" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system global' \n 'set ssl-static-key-ciphers disable '."),

        ("System Settings", "Activer le cryptage global fort", 
         "Activer FortiOS pour n'utiliser que le chiffrement fort et permettre uniquement l'utilisation de chiffres forts pour la communication.", 
         "L'activation du chiffrement fort assure une protection accrue des communications, réduisant ainsi le risque de compromission des données en transit.", 
         "conforme" if any("set strong-crypto enable" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system global' \n 'set strong-crypto enable'."),

        ("System Settings", "L'interface graphique de gestion écoute la version sécurisée de TLS", 
         "l'accès à l'interface graphique d'administration (GUI) est sécurisé en écoutant sur la version TLS sécurisée.", 
         " prévenir les attaques de type Man-in-the-Middle (MiTM).", 
         "conforme" if any("set admin-https-ssl-versions tlsv1-3" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'config system global' \n 'set admin-https-ssl-versions tlsv1-3'."),

        ("System Settings", "Activer 'Password Policy'", 
         "Il est important d'utiliser des mots de passe sécurisés et complexes pour prévenir l'accès non autorisé au dispositif FortiGate.", 
         "Les mots de passe faibles peuvent être facilement découverts par les pirates informatiques, ce qui conduit à un accès non autorisé à FortiGate. En fonction du privilège d'accès du compte compromis, l'attaquant peut modifier des paramètres importants.", 
         "conforme" if any("config system password-policy" in line and "set status enable" in line and "set apply-to admin-password ipsec-preshared-key" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system password-policy ' \n 'set status enable ' \n 'set apply-to admin-password ipsec-preshared-key'."),
        
        ("System Settings", "les tentatives de mot de passe de l'administrateur", 
         "Limiter le nombre de tentatives de connexion échouées pour les administrateurs", 
         "Les attaquants continueront d'essayer d'accéder à l'appareil par des attaques par force brute sans interruption, ce qui peut conduire à une connexion réussie.", 
         "conforme" if any("set admin-lockout-threshold 3" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system global' \n 'set admin-lockout-threshold 3 '."),

        ("System Settings", "le temps de verrouillage de l'administrateur ", 
         "définir une durée de verrouillage après le dépassement de ce seuil pour les administrateurs .", 
         "Les attaquants continueront d'essayer d'accéder à l'appareil par des attaques par force brute sans interruption, ce qui peut conduire à une connexion réussie.", 
         "conforme" if any("set admin-lockout-duration 900" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system global' \n 'set admin-lockout-duration 900."),
        
        ("System Settings", "le service SNMPv3 est activé ", 
         "S'assurer que seul le service SNMPv3 est activé et que SNMPv1 et SNMPv2c sont désactivés.", 
         "Certains serveurs SNMP plus anciens qui n'exécutent que SNMPv1 ou SNMPv2c ne pourront pas interroger ce pare-feu.", 
         "conforme" if any("config system snmp sysinfo\\s+set status enable" in line for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system snmp sysinfo' \n 'set status enable '."),

        ("System Settings", "le délai d'inactivité", 
         "La période d'inactivité est le laps de temps pendant lequel un administrateur reste connecté à l'interface graphique sans aucune activité.", 
         "Cela vise à empêcher quelqu'un d'accéder au FortiGate si le PC de gestion est laissé sans surveillance.", 
         "conforme" if any("set admintimeout 5" in line for line in output) else "Non conforme", 
         "Faible", 
         "Configurer la commande 'config system global' \n 'set admintimeout 5'."),
        
        ("System Settings", "les canaux d'accès cryptés sont activés", 
         "Autoriser uniquement l'accès HTTPS à l'interface graphique et l'accès SSH à l'interface CLI.", 
         "En n'autorisant que l'accès chiffré, nous rendons plus difficile l'utilisation d'attaques de type Man-in-the-Middle (MiTM) pour intercepter les informations d'identification de connexion.", 
         "conforme" if any(re.search(r"set allowaccess.*(ssh|https|ping)", line) for line in output) else "Non conforme", 
         "Élevée", 
         "Configurer la commande 'config system interface' \n 'edit port1'`\n 'set allowaccess ssh https ping '."),
        
        ("System Settings", "Modifier les ports d'administration par défaut", 
         "Autoriser uniquement l'accès HTTPS à l'interface graphique et l'accès SSH à l'interface CLI.", 
         "Augmenter la sécurité des ports d'administration de FortiGate, les changer à partir des ports par défaut réduira la surface d'attaque en cas de ciblage de l'accès administratif de FortiGate. ", 
         "Non conforme" if any(" set admin-port 80" in line and "set admin-sport 443 "in line for line in output) else "conforme", 
         "Moyenne", 
         "Configurer la commande 'config system global' \n 'set admin-https-redirect disable'`\n 'set admin-port <uncommon port>'\n 'set admin-server-cert 'self-sign' ' \n 'set admin-sport <uncommon port>'."),
        
        ("Policy and Objects", "désactiver les politiques inutilisées  ", 
         "Les règles inutilisées doivent être désactivées et consignées. Recommandation de les examiner deux fois par an ou en ligne avec les pratiques du PBC .", 
         "Protection des données sensibles contre les menaces en ligne .", 
         " conforme" if any(" firewall iprope show 100004 32" in line and "idx=2 pkts/bytes=144967/135758174 asic_pkts/asic_bytes=0/0 flag=0x0 hit count:663 "in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer la commande 'diag firewall iprope clear 100004 32'."),

        ("Policy and Objects", "les politiques n'utilisent pas 'ALL'   ", 
         " Assurez-vous que toutes les politiques de sécurité en vigueur précisent clairement les protocoles / services qu'elles autorisent .", 
         "réduire les risques, optimiser les performances , faciliter la gestion du réseau .", 
         "Conforme" if any("set service HTTPS" in line and "set service HTTP" in line and "set service FTP" in line and "set service SNMP" in line for line in output) else "Non conforme",
         "Moyenne", 
         "Configurer la commande 'set service FTP' et 'set service SNMP'."),

        ("Policy and Objects", "Assurez-vous qu'une politique de pare-feu refuse tout le trafic vers/depuis les adresses IP associées à Tor, ainsi que les serveurs malveillants.", 
         "Les règles de pare-feu doivent inclure une règle de refus pour le trafic entrant/sortant des adresses IP Tor, des serveurs malveillants ou des scanners en utilisant la base de données ISDB (Internet Service Database).", 
         "Cela vise à protéger le réseau contre les attaques, à optimiser les performances et les ressources réseau, ainsi qu'à simplifier la gestion continue de la sécurité.", 
         "Conforme" if any("set service 'Tor'" in line or "set service 'Malicious-Server'" in line or "set service 'Scanner-IP'" in line for line in output) else "Non conforme", 
         "Moyenne", 
         "Configurer les commandes 'set service 'Tor'', 'set service 'Malicious-Server'' et 'set service 'Scanner-IP''."),
        

        ("Policy and Objects", "Assurez-vous que la journalisation est activée sur toutes les politiques de pare-feu.", 
         "La journalisation doit être activée pour toutes les politiques de pare-feu, y compris la politique implicite de refus par défaut.", 
         "Cela permet aux analystes de la SOC de mener des enquêtes approfondies sur les incidents de sécurité, en particulier pour les activités de recherche de menaces ou de réponse aux incidents.", 
         "conforme" if not any("set logtraffic all" in line and ("set action accept" in line or "set action deny" in line) for line in output) else "Non Conforme", 
         "Moyenne", 
         "Configurer les commandes 'set action accept' et 'set logtraffic all'."),
        
        ("Security Profiles", "Appliquer un profil de sécurité IPS aux stratégies de pare-feu.",
         "Assurez-vous que tout le trafic traversant entre les réseaux sur le FortiGate est inspecté par un profil de sécurité IPS.",
         "Cela vise à réduire les risques de compromission, à renforcer la conformité réglementaire et à améliorer la visibilité et la gestion des menaces.",
         "Conforme" if any("set ips-sensor" in line and "set action accept" in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande 'set ips-sensor <nom_du_profil>' pour tous les profils de sécurité IPS."),
        
        ("Security Profiles", "Assurez-vous que les mises à jour automatiques des définitions antivirus sont configurées.",
         "Veillez à ce que le FortiGate soit configuré pour accepter les mises à jour automatiques des définitions antivirus.",
         "Cela garantit une protection contre les menaces de logiciels malveillants, réduit le risque d'infections et de propagation des logiciels malveillants, et permet une réponse plus rapide aux nouvelles menaces.",
         "Conforme" if any("config system autoupdate schedule" in line and "set frequency automatic" in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande 'config system autoupdate schedule'\n 'set status enable'\n 'set frequency automatic'."),

        ("Security Profiles", "Appliquer le profil de sécurité antivirus aux politiques.",
         "S'assurer que le trafic traversant entre les réseaux sur le FortiGate est inspecté par un profil de sécurité antivirus.",
         "Cela garantit une protection contre les menaces de logiciels malveillants, réduit le risque d'infections et de propagation des logiciels malveillants, et permet une réponse plus rapide aux nouvelles menaces.",
         "Conforme" if any("set profile-type antivirus" in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande 'set profile-type antivirus'\n 'set profile 'Antivirus_Profile_Name''."),

        ("Security Profiles", "Activer la base de données de prévention des épidémies de FortiGuard.",
         "Assurer que l'inspection antivirus du FortiGate utilise la base de données de prévention des épidémies en plus de la détection basée sur les signatures antivirus.",
         "Réduire le temps de réponse aux incidents.",
         "Conforme" if any("set outbreak-prevention block" in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande 'set outbreak-prevention block'."),

        ("Security Profiles", "Activer la détection de logiciels malveillants basée sur l'IA / heuristique.",
         "La détection basée sur l'IA / l'heuristique doit être activée",
         "une protection avancée contre les menaces inconnues, les attaques zero-day et les tactiques d'évasion sophistiquées utilisées par les cybercriminels.",
         "Conforme" if any("set machine-learning-detection enable" in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande 'config antivirus settings','set machine-learning-detection enable'."),
        
        ("Security Profiles", "Activer la détection des logiciels indésirables (grayware) sur l'antivirus.",
          "La détection des logiciels indésirables (grayware) devrait être activée.",
          "identifier et en supprimer les logiciels potentiellement dangereux, tout en améliorant les performances et l'expérience utilisateur.",
          "Conforme" if any("set grayware enable" in line for line in output) else "Non conforme",
          "Moyenne",
          "Configurer la commande 'set grayware enable'."),

        
        ("Security Profiles", "Assurer que l'analyse en ligne avec le service de bac à sable basé sur l'IA de FortiGuard est activée .",
         "La numérisation en ligne est prise en charge lorsque le FortiGate est sous licence avec le service de bac à sable basé sur l'IA de FortiGuard (FAIS). ",
         " la détection basée sur les signatures antivirus existantes et agit également comme une couche de défense supplémentaire sur la fonctionnalité heuristique de l'AV FortiGate.",
         "Conforme" if any("config system fortiguard" in line and "set sandbox-region 'Global'"in line and "set sandbox-inline-scan enable" in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande 'set sandbox-region 'Global' '\n 'set sandbox-inline-scan enable."),

        ("Security Profiles", "Activer le filtre DNS de blocage de domaine de commande et contrôle des botnets .",
         "Activer le blocage des domaines de commande et contrôle des botnets pour bloquer l'accès des botnets au stade de la résolution DNS. ",
         " Bloquer l'accès aux sites Web des botnets au stade de la résolution DNS fournit une couche de défense supplémentaire.",
         "Conforme" if any("set botnet-domains enable " in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande 'set botnet-domains enable' ."),

        ("Security Profiles", "Activer la journalisation de toutes les requêtes et réponses DNS du filtre DNS .",
         "Le filtre DNS doit enregistrer toutes les requêtes et réponses DNS. ",
         " permet aux SOC ou aux analystes de sécurité de mener des investigations supplémentaires sur les incidents de sécurité, notamment dans le cadre de la chasse aux menaces ou des activités de réponse aux incidents.",
         "Conforme" if any("set log-all-domain enable" in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande 'set log-all-domain enable' ."),

        ("Security Profiles", "Appliquer le profil de sécurité du filtre DNS aux politiques.",
         "S'assurer que le trafic traversant vers Internet sur le FortiGate a un profil de sécurité du filtre DNS l'inspectant. ",
         " prévenir les attaques basées sur le DNS, filtrer le contenu indésirable et non autorisé, protéger contre les fuites d'informations.",
         "Conforme" if any("set utm-status enable" in line and "set utm-profile" in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande 'set utm-profile <nom_du_profil_DNS_Filter>' ."),

        ("Security Profiles", "Bloquer les catégories à haut risque sur le contrôle des applications.",
         "Assurer que le contrôle des applications FortiGate bloque les applications à haut risque pour réduire la surface d'attaque. ",
         "  réduire les risques pour les données et les systèmes ..",
         "Conforme" if any("config firewall profile-protocol-options" in line and "set filter p2p disable" in line and "set filter proxy disable " in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande 'config firewall profile-protocol-options','set filter p2p disable','set filter proxy disable'."),

        ("Security Profiles", "Bloquer les applications s'exécutant sur des ports non par défaut.",
         "S'assurer que le Contrôle des Applications du FortiGate bloque les applications s'exécutant sur des ports non par défaut. ",
         "réduire les risques d'attaques et de fuites de données.",
         "Conforme" if any("set enforce-default-app-port enable" in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande ' config application list','edit <profile name>','set enforce-default-app-port enable'."),

        ("Security Profiles", "S'assurer que tout le trafic lié au Contrôle des Applications est journalisé .",
         "S'assurer qu'aucune catégorie n'est définie comme 'Autoriser' dans le Contrôle des Applications du FortiGate.",
         "renforce la détection et la réponse aux incidents de sécurité et facilite les enquêtes post-incidents et forensiques.",
         "Conforme" if any("config system syslog" in line and " set log-application-control enable" in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande ' config system syslog ','set log-application-control enable'."),

        ("Security Profiles", "Appliquer le profil de sécurité du Contrôle des Applications aux politiques.",
         "S'assurer que le trafic traversant entre les réseaux sur le FortiGate est inspecté par un profil de sécurité du Contrôle des Applications.",
         "Contrôle granulaire des applications , Prévention des menaces avancées.",
         "Conforme" if any("set utm-status enable" in line and " set utm-profiles" in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande ' config firewall policy ','set utm-status enable','set utm-profile 'nom_du_profil_du_contrôle_des_applications''."),

        ("VPN", "Appliquer un certificat signé de confiance pour le portail VPN.",
         "Appliquer un certificat signé d'une autorité de certification (CA) de confiance au portail VPN SSL pour permettre aux utilisateurs de se connecter en toute sécurité et avec confiance.",
         "Renforcement de la sécurité des communications , Authentification du serveur VPN , Réduction des risques d'attaques .",
         "Conforme" if any("config vpn ssl settings" in line and " set servercert " in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande ' config vpn ssl settings ','set servercert <nom_du_certificat>'."),

        ("VPN", "Activer les versions limitées de TLS pour SSL VPN",
         "Activer et désactiver les versions de TLS et les suites de chiffrement pour un contrôle plus granulaire des connexions SSL VPN et l'application de connexions plus sécurisées.",
         "Protection contre les attaques par downgrade , Gestion simplifiée des certificats.",
         "Conforme" if any("config vpn ssl settings" in line and "set ssl-max-prot-ver tls1-3" in line and " set ssl-min-proto-ver tls1-2 " in line and "set algorithm high" in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande ' config vpn ssl settings ','set ssl-max-prot-ver tls1-3', 'set ssl-min-proto-ver tls1-2', 'set algorithm high'."),
        
        ("Logs and Reports", "Activer la journalisation des événements",
         "Activer la journalisation des événements pour permettre la génération et la révision des journaux.",
         "Surveillance en temps réel et alertes, Support aux enquêtes post-incident , Support pour les investigations légales.",
         "Conforme" if any("config log eventfilter" in line and "set event enable" in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande 'config log eventfilter ','set event enable' ."),

        ("Logs and Reports", "Chiffrer la transmission des journaux vers FortiAnalyzer / FortiManager ",
         "Activer le chiffrement des journaux envoyés à FortiAnalyzer ou FortiManager.",
         "protège contre les attaques MitM, augmente la fiabilité et la confiance dans les analyses de sécurité, et protège les informations sensibles.",
         "Conforme" if any("config log fortianalyzer setting" in line and "set reliable enable" in line and "set enc-algorithm high" in line for line in output) else "Non conforme",
         "Élevée",
         "Configurer la commande 'config log fortianalyzer setting','set event enable' , 'set enc-algorithm high' ."),

        ("Logs and Reports", "Journalisation et reporting centralisés",
         "Les journaux de l'appareil doivent être envoyés à un appareil centralisé pour la collecte, la conservation et le reporting des journaux. Cela pourrait être un SIEM, un périphérique syslog, FortiAnalyzer, FortiManager, etc.",
         "la détection des menaces, la réponse aux incidents, la gestion des opérations, la conformité réglementaire et simplifie l'analyse et le reporting de sécurité.",
         "Conforme" if any("config log syslogd2 setting" in line and "set status enable" in line and "set server " in line for line in output) else "Non conforme",
         "Moyenne",
         "Configurer la commande 'config log fortianalyzer setting','set event enable' , 'set server <adresse_IP_du_serveur>' ."),

    ]

    
    data.extend(checkpoints)

def generate_drawingsF(data):
    """Générer des graphiques basés sur les données traitées."""
    print("#####################generate_drawingsF(data)")
    # Graphique à barres
    categories = [d[0] for d in data]
    errors = [d.count("Non conforme") for d in data]

   

    # Graphique circulaire
    labels = list(set(categories))
    sizes = [categories.count(label) for label in labels]
    plt.figure(figsize=(8, 6))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%')
    plt.axis('equal')
    plt.title('Répartition des erreurs par catégorie')
    plt.savefig("dessin_rapport_audit_fortinet.png")

def create_excelF(data):
    """Créer un fichier Excel contenant les données du rapport d'audit."""
    print("#########################create_excelF(data)")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"rapport_audit_forti_{timestamp}.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rapport Audit"

    # En-têtes de colonne
    headers = ["Catégorie", "Point de contrôle", "Description", "Impact", "Conformité", "Criticité", "Remédiation"]
    sheet.append(headers)

    # Largeurs de colonne
    column_widths = [20, 20, 50, 50, 15, 15, 50]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Style de l'en-tête
    header_fill = PatternFill(start_color="EC407A", end_color="EC407A", fill_type="solid")
    header_font = Font(color="060605", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Écrire les données dans la feuille de calcul
    for row in data:
        sheet.append(row)

    # Parcourir chaque cellule pour trouver "Network Settings" et la colorer en bleu
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Network Settings" in str(cell.value):
                cell.fill = PatternFill(start_color="2F9AF8", end_color="2F9AF8", fill_type="solid")  # Bleu
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")  # Couleur de la police en blanc pour une meilleure visibilité


    # Parcourir chaque cellule pour trouver "System Settings" et 
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "System Settings" in str(cell.value):
                cell.fill = PatternFill(start_color="f97434", end_color="f97434", fill_type="solid")  
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")  # Couleur de la police en blanc pour une meilleure visibilité
    
    # Parcourir chaque cellule pour trouver "Policy and Objects" et
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Policy and Objects" in str(cell.value):
                cell.fill = PatternFill(start_color="87228b", end_color="87228b", fill_type="solid") 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")

    
    # Parcourir chaque cellule pour trouver "Security Profiles" et
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Security Profiles" in str(cell.value):
                cell.fill = PatternFill(start_color="9e87ee", end_color="9e87ee", fill_type="solid") 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")

    
    # Parcourir chaque cellule pour trouver "Security Fabric" et
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Security Fabric" in str(cell.value):
                cell.fill = PatternFill(start_color="1BF246", end_color="1BF246", fill_type="solid") 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")

    # Parcourir chaque cellule de la première colonne pour trouver "VPN" et 1D8348
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if "VPN" in str(cell.value):
                cell.fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")



     # Parcourir chaque cellule pour trouver "Logs and Reports" et
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if "Logs and Reports" in str(cell.value):
                cell.fill = PatternFill(start_color="DB1BF2", end_color="DB1BF2", fill_type="solid") 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(color="FFFFFF")


     # Styles conditionnels pour la conformité
    for row in sheet.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column == 5:  # Colonne de conformité
                if cell.value.lower() == "conforme":
                   cell.fill = PatternFill(start_color="08A91E", end_color="08A91E", fill_type="solid")  # Vert 
                elif cell.value.lower() == "non conforme":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rouge
                cell.font = Font(color="FFFFFF")  # Texte en blanc pour toutes les cellules colorées
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



    # Styles conditionnels pour la colonne Criticité 
    for row in sheet.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column == 6:  # Colonne 6
                if "Élevée" in str(cell.value):
                    cell.font = Font(color="FF0000")  # Rouge
                elif "Moyenne" in str(cell.value):
                    cell.font = Font(color="FFA500")  # Orange
                elif "Faible" in str(cell.value):
                    cell.font = Font(color="00FF00")  # Vert
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)      
     


    # Ajuster la hauteur des lignes pour s'adapter au contenu
    adjust_row_heightF(sheet)

    workbook.save(filename=excel_filename)

def adjust_row_heightF(sheet):
    """Ajuster automatiquement la hauteur des lignes pour s'adapter au contenu."""
    print("#######################adjust_row_heightF(sheet)")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_height = 0
        for cell in row:
            if cell.value:
                cell_value = str(cell.value)
                lines = cell_value.count('\n') + 1
                max_line_length = max(len(line) for line in cell_value.split('\n'))
                estimated_height = (lines + max_line_length // sheet.column_dimensions[cell.column_letter].width) * 15
                max_height = max(max_height, estimated_height)
        sheet.row_dimensions[row[0].row].height = max_height
        
@app.route('/hardeningF/<ip>/<username>/<password>', methods=['POST'])
def hardeninF(ip, username, password):
    while True:
        output = ssh_connectF(ip, username, password)
        if output:
            data = process_outputF(output)
            create_excelF(data)
            generate_drawingsF(data)
            return jsonify({"message": "forti aud"})
        else:
            print("Échec de connexion SSH. Veuillez réessayer.")
            return jsonify({"message": "forti_aud"}), 500

if __name__ == '__main__':
    app.run(debug=True)







